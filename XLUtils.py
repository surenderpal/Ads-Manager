import openpyxl

def getRowCounnt(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnum,data):
    wb=openpyxl.load_workbook(file)
    sheet=wb[sheetName]
    sheet.cell(row=rownum,column=columnum).value=data
    wb.save(file)
    