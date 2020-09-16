from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium .webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager #for chrome
from webdriver_manager.firefox import GeckoDriverManager #for firefox
from webdriver_manager.microsoft import IEDriverManager #for IE
from webdriver_manager.microsoft import EdgeChromiumDriverManager #for edge
from time import sleep


# driver=webdriver.Chrome()
# driver.get('https://location-software-np.groundtruth.com/')
# time.sleep(2)
# # driver.execute_script('window.scrollBy(0,5000),""') # scroll by pixel
# # driver.execute_script('window.scrollBy(0,document.body.scrollHeight)') #scroll till the bottom of the page
# # element=driver.find_element_by_xpath('//*[@id="lifenstylewidget"]') #scroll till the element is found
# # # element=driver.find_element_by_link_text("About us ")
# # driver.execute_script("arguments[0].scrollIntoView();", element)

# # driver.execute_script("arguments[0].scrollIntoView();", element)
# # time.sleep(2)
# # driver.close()
# def login(Username,Password):
#     driver.find_element_by_name('username').send_keys(Username)
#     driver.find_element_by_name('password').send_keys(Password)
#     driver.find_element_by_id('submit').click()

# # login('surender.pal@groundtruth.com','Surenderpal@1991')
# time.sleep(2)
# print('Login successfully in location software')
# driver.find_element_by_xpath("//div[@class='hamburger-icon']").click() #hamburger menu

# i=driver.find_element_by_xpath("//a[@class='info-circle-button']")
# brands=driver.find_element_by_xpath("//span[contains(text(),'Brands')]") 
# LG=driver.find_element_by_xpath("//span[contains(text(),'Location Groups')]")
# actions=ActionChains(driver)
# actions.move_to_element(brands).move_to_element(LG).click().perform()
# am=driver.find_element_by_xpath("//button[contains(text(),'Audience Manager')]")
# lm=driver.find_element_by_xpath("//button[contains(text(),'Location Manager ')]")
# time.sleep(2)
# info=driver.find_element_by_xpath("//div[@class='dropdown-info-container']/div[@class='info-container']")
# # info.click()
# actions.move_to_element(am).move_to_element(lm).click().perform() #.move_to_element(info)
# print('clicked on info')

# #----hover pratice-----
# driver.get('https://yomovies.link/')
# driver.maximize_window()
# genre=driver.find_element_by_xpath("//li/a[(contains(text(), 'Genre'))]")
# action=driver.find_element_by_xpath("//li/a[(contains(text(), 'Action'))]")
# biography=driver.find_element_by_xpath("//li/a[(contains(text(), 'Biography'))]")
# music=driver.find_element_by_xpath("//li/a[(contains(text(), 'Music'))]")
# fantasy=driver.find_element_by_xpath("//li/a[(contains(text(), 'Fantasy'))]")
# Fiction=driver.find_element_by_xpath("//li/a[(contains(text(), 'Science Fiction'))]")
# Adventure=driver.find_element_by_xpath("//li/a[(contains(text(), 'Adventure'))]")
# Comedy=driver.find_element_by_xpath("//li/a[(contains(text(), 'Comedy'))]")
# War=driver.find_element_by_xpath("//li/a[(contains(text(), 'War'))]")
# actions=ActionChains(driver)
# actions.move_to_element(genre).move_to_element(action).move_to_element(biography).move_to_element(music).move_to_element(fantasy).move_to_element(Fiction).move_to_element(Adventure).move_to_element(Comedy).move_to_element(War).click().perform()
# print('learnt hover')

##---double click pratice ------- 

# driver.get("https://testautomationpractice.blogspot.com/")
# driver.maximize_window()
# element=driver.find_element_by_xpath("//button[contains(text(), 'Copy Text')]")
# actions=ActionChains(driver)
# actions.double_click(element).perform()  #double click
# print('double clicked on element')

##-----pratice right click-----

# driver.get('https://swisnl.github.io/jQuery-contextMenu/demo.html')
# time.sleep(3)
# element=driver.find_element_by_xpath("//span[contains(text(), 'right click me')]")
# actions=ActionChains(driver)
# actions.context_click(element).perform()
# print('Right clicke performed')


##---Drag and drop-----

# driver.get('http://dhtmlgoodies.com/scripts/drag-drop-custom/demo-drag-drop-3.html')
# time.sleep(2)
# source=driver.find_element_by_id("box6")
# target=driver.find_element_by_id("box106")
# actions=ActionChains(driver)
# actions.drag_and_drop(source,target).perform()
# print('Dragged and dropped!!!')


# # driver.maximize_window()
# time.sleep(2)
# trash=driver.find_element_by_id('slider')
# # s=driver.find_element_by_id("draggable")
# s=driver.find_element_by_xpath("//*[@alt='the peaks of high tatras']")

# # t=driver.find_element_by_id("droppable")
# t=driver.find_element_by_id("trash")

# # driver.execute_script("arguments[0].scrollIntoView();",trash) #scroll page
# actions=ActionChains(driver)
# # actions.drag_and_drop(s,t).perform()
# actions.move_to_element(trash).perform()
# actions.drag_and_drop(s,t).perform()
# print('Drag and dropped')


# driver.get("https://testautomationpractice.blogspot.com/")
# time.sleep(2)
# trash=driver.find_element_by_id('slider')
# s=driver.find_element_by_xpath("//*[@alt='the peaks of high tatras']")
# t=driver.find_element_by_id("trash")
# actions=ActionChains(driver)
# actions.move_to_element(trash).perform()
# actions.drag_and_drop(s,t).perform()
# print('Drag and dropped')


##---how to upload the file------
# driver.get('https://testautomationpractice.blogspot.com/')
# time.sleep(5)
# driver.maximize_window()
# scroll=driver.find_element_by_id("HTML7")
# # driver.execute_script("arguments[0].scrollIntoView();",scroll)
# actions=ActionChains(driver)
# actions.move_to_element(scroll).perform()
# time.sleep(2)

# # actions=ActionChains(driver)
# # actions.move_to_element(scroll).perform()
# driver.switch_to.frame(0)
# driver.find_element_by_id("RESULT_FileUpload-10").send_keys("/Users/surenderpal/Downloads/sukhoi.jpg") #frame_Id=FSForm
# # /Users/surenderpal/Downloads/surender.png

##2nd pratice-----

# driver.get('https://filebin.net/')
# time.sleep(4)
# driver.switch_to.frame(1)
# print('Switched to frame!!')
# driver.find_element_by_id("fileField").send_keys('/Users/surenderpal/Downloads/sukhoi.jpg')

##---3rd pratice-----

# driver.get('https://safenote.co/upload-file')
# time.sleep(4)
# driver.switch_to.frame(driver.find_element_by_id("dropzoneFrom"))
# actions=ActionChains(driver)
# scroll=driver.find_element_by_xpath("//*[@class='jumbotron']")
# actions.move_to_element(scroll).perform()
# print('scrolling done!!!')
# time.sleep(4)
# driver.find_element_by_id('dropzoneFrom').send_keys('/Users/surenderpal/Downloads/sukhoi.jpg')

# ##---4th pratice-----
# driver.get("https://gofile.io/uploadFiles")
# time.sleep(4)
# driver.find_element_by_xpath("//button[@class='btn btn-primary btn-lg']").send_keys('/Users/surenderpal/Downloads/VID-20200819-WA0006.mp4')
# print('file uploaded successfully!!!')


##--read data from excel----

# import openpyxl
# path="/Users/surenderpal/Downloads/DS 160.xlsx"
# wb=openpyxl.load_workbook(path)
# sheet=wb["Info. for DS160"]
# rows=sheet.max_row
# cols=sheet.max_column
# print(rows)
# print(cols)
# for r in range(1,rows+1):
#     for c in range(1, cols):
#         print(sheet.cell(row=r,column=c).value, end="       ")
#     print()

## pratice --2 _------

# import openpyxl
# path='/Users/surenderpal/Downloads/Geotarget.xlsx'
# wb=openpyxl.load_workbook(path)
# sheet=wb['rahul']
# for r in range(1, 6):
#     for c in range(1,4):
#         sheet.cell(row=r,column=c).value='rahul'
# wb.save(path)

# pratice for links:
# def browserName(bname):#bname
bname='chrome'
if bname=='chrome':
    driver=webdriver.Chrome(ChromeDriverManager().install()) #this will install latest chrome driver manager
elif bname=='firefox':
    driver=webdriver.Firefox(executable_path=GeckoDriverManager().install()) #it will install latest firefox driver manager
elif bname=='ie':
    driver=webdriver.Ie(IEDriverManager().install())
elif bname=='edge':
    driver=webdriver.Edge(EdgeChromiumDriverManager().install())
elif bname=='safari':
    driver=webdriver.Safari()
else:
    print('Please pass the correct browser name' + bname)
driver.get('https://ads-release-3-14-np.groundtruth.com/')
# browserName('firefox')


# login class to automate the login process
class login():
    driver.maximize_window()
    def AdManager(self,username,password):
        driver.find_element_by_name('username').send_keys(username)
        driver.find_element_by_name('password').send_keys(password)
        driver.find_element_by_id('btn-signin-signIn').click()

def links():
    links=driver.find_elements(By.TAG_NAME, "a")
    print('total links on page',len(links))
    for link in links:
        print(link.text)
    

links()
l=login()
l.AdManager('surender.pal@groundtruth.com','Surenderpal@1991')
sleep(5)
